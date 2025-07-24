from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import xml.etree.ElementTree as ET
import os
from api.ds_i_utils import d_si
from api.ds_i_utils import convert_unit
import base64
import tempfile
from openpyxl.drawing.image import Image
from io import BytesIO
import logging
import mimetypes
import base64
import tempfile
import logging
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image as PILImage
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image as ExcelImage


# Function to handle None values in XML parsing
def gt(el):
    return el.text if el is not None and el.text is not None else "-"

# Border style for table cells
border = Border(
    left=Side(border_style='thin', color='000000'),
    right=Side(border_style='thin', color='000000'),
    top=Side(border_style='thin', color='000000'),
    bottom=Side(border_style='thin', color='000000')
)

# Function to apply border to cells in a given range
def apply_borders(ws, start_row, end_row, start_col, end_col):
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.border = border

# Utility function to create Excel table headers
def create_table_header(ws, row, headers):
    col = 1
    for header in headers:
        ws.cell(row=row, column=col, value=header).font = Font(bold=True)
        ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
        col += 1
    return row + 1

# Fungsi pembantu untuk menyimpan base64 ke file gambar sementara
def save_base64_image(base64_str, file_extension=".png"):
    try:
        # Decode base64
        image_data = base64.b64decode(base64_str)
        
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=file_extension)
        temp_file.write(image_data)
        temp_file.close()
        
        return temp_file.name
    except Exception as e:
        logging.error(f"Error saving base64 image: {str(e)}")
        return None

#RUMUS
def render_latex_to_excel_image(latex_expr: str) -> BytesIO:
    fig = plt.figure(figsize=(0.01, 0.01), dpi=200)
    fig.text(0.1, 0.5, f"${latex_expr}$", fontsize=14)
    plt.axis('off')
    buf = BytesIO()
    fig.savefig(buf, format='png', bbox_inches='tight', dpi=200)
    buf.seek(0)
    plt.close(fig)
    return buf

#GAMBAR
def mime_to_extension(mime_type: str) -> str:
    if not mime_type:
        return ".png"
    ext = mimetypes.guess_extension(mime_type)
    if ext:
        return ext
    fallback = {
        "image/jpeg": ".jpg",
        "image/jpg": ".jpg",
        "image/png": ".png",
        "image/gif": ".gif",
        "image/bmp": ".bmp",
        "image/tiff": ".tiff"
    }
    return fallback.get(mime_type.lower(), ".png")

def add_image_to_worksheet(ws, base64_str, mime_type, caption, row, col='E'):
    try:
        base64_str = base64_str.replace('\n', '').replace('\r', '').replace(' ', '')
        if base64_str.startswith('data:'):
            base64_str = base64_str.split(',', 1)[1]

        try:
            image_data = base64.b64decode(base64_str)
        except Exception as e:
            logging.error(f"[Base64 Decode Error] Tidak bisa decode: {e}")
            ws[f'{col}{row}'] = "[Gambar gagal decode]"
            return 2

        extension = mime_to_extension(mime_type)

        with tempfile.NamedTemporaryFile(delete=False, suffix=extension) as tmp_file:
            tmp_file.write(image_data)
            tmp_file_path = tmp_file.name

        try:
            with PILImage.open(tmp_file_path) as pil_img:
                pil_img.load()
                width, height = pil_img.size
        except Exception as e:
            logging.error(f"[Gambar Invalid] {e}")
            ws[f'{col}{row}'] = "[Gambar tidak valid]"
            os.unlink(tmp_file_path)
            return 2

        img = ExcelImage(tmp_file_path)
        max_width = 200
        if img.width > max_width:
            ratio = max_width / img.width
            img.width = max_width
            img.height = int(img.height * ratio)

        img_anchor = f'{col}{row}'
        ws.add_image(img, img_anchor)

        total_rows = 5  

        if caption:
            ws[f'{col}{row + 4}'] = f"Keterangan: {caption}"
            total_rows += 1

        logging.info(f"Gambar berhasil ditambahkan di {img_anchor}")
        return total_rows

    except Exception as e:
        logging.error(f"[Gagal Menambahkan Gambar] {e}")
        ws[f'{col}{row}'] = "[Gambar gagal dimuat]"
        return 2



# Function to convert XML to Excel
def convert_xml_to_excel(xml_file_path: str):
    try:
        
        
        # Load XML using ElementTree
        tree = ET.parse(xml_file_path)
        root = tree.getroot()

        # Initialize Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Hasil"
        row = 1

        ns = {  # XML namespaces
            'dcc': 'https://ptb.de/dcc',
            'si': 'https://ptb.de/si'
        }

        # SOFTWARE
        software = root.find('.//dcc:software', ns)
        software_name = gt(software.find('dcc:name/dcc:content', ns)) if software else "N/A"
        version = gt(software.find('dcc:release', ns))

        ws.cell(row=row, column=1, value="Perangkat Lunak").font = Font(bold=True, size=14)
        row += 2
        ws.cell(row=row, column=1, value="Nama")
        ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=5, value=software_name)
        row += 2
        ws.cell(row=row, column=1, value="Versi")
        ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=5, value=version)
        row += 2

        # ADMINISTRATIVE DATA
        administrasi = root.find('.//dcc:coreData', ns)
        country_calib = gt(administrasi.find('dcc:countryCodeISO3166_1', ns))
        place = gt(administrasi.find('dcc:performanceLocation', ns))
        #bahasa
        langs = administrasi.findall('dcc:usedLangCodeISO639_1', ns)
        used_lang = ', '.join(lang.text for lang in langs if lang is not None and lang.text)
        mandatory_langs = administrasi.findall('dcc:mandatoryLangCodeISO639_1', ns)
        mandatory_lang = ', '.join(lang.text for lang in mandatory_langs if lang is not None and lang.text)
        
        order = gt(administrasi.find('dcc:identifications/dcc:identification/dcc:value', ns))
        order_issuer = gt(administrasi.find('dcc:identifications/dcc:identification/dcc:issuer', ns))
        sertifikat = gt(administrasi.find('dcc:uniqueIdentifier', ns))

        ws.cell(row=row, column=1, value="Data Administrasi").font = Font(bold=True, size=14)
        row += 2
        ws.cell(row=row, column=1, value="Negara tempat kalibrasi")
        ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=5, value=country_calib)
        row += 2
        ws.cell(row=row, column=1, value="Tempat kalibrasi")
        ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=5, value=place)
        row += 2
        ws.cell(row=row, column=1, value="Bahasa yang digunakan")
        ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=5, value=used_lang)
        row += 2
        ws.cell(row=row, column=1, value="Bahasa yang diwajibkan")
        ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=5, value=mandatory_lang)
        row += 2
        ws.cell(row=row, column=1, value="Nomor order")
        ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=5, value=order)
        row += 2
        ws.cell(row=row, column=1, value="Penerbit nomor order")
        ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=5, value=order_issuer)
        row += 2
        ws.cell(row=row, column=1, value="Nomor sertifikat")
        ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=5, value=sertifikat)
        row += 2


        # TIMELINE
        begin_date = gt(administrasi.find('dcc:beginPerformanceDate', ns))
        end_date = gt(administrasi.find('dcc:endPerformanceDate', ns))
        issue_date = gt(administrasi.find('dcc:issueDate', ns))

        ws.cell(row=row, column=1, value="Linimasa Pengukuran").font = Font(bold=True, size=14)
        row += 2
        ws.cell(row=row, column=1, value="Tanggal mulai pengukuran")
        ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=5, value=begin_date)
        row += 2
        ws.cell(row=row, column=1, value="Tanggal akhir pengukuran")
        ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=5, value=end_date)
        row += 2
        ws.cell(row=row, column=1, value="Tanggal pengesahan")
        ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=5, value=issue_date)
        row += 2
        
        
        #OBJECTS
        objects = root.findall('.//dcc:items/dcc:item', ns)
        ws.cell(row=row, column=1, value="Deskripsi Objek yang Dikalibrasi/Diukur").font = Font(bold=True, size=14)
        row += 2

        # Iterate over each object
        for i, objek in enumerate(objects, start=1):
            
            jenis_elements = objek.findall('dcc:name/dcc:content', ns)
            jenis = ', '.join([gt(jenis_element) for jenis_element in jenis_elements if jenis_elements])

            merek = gt(objek.find('.//dcc:manufacturer/dcc:name/dcc:content', ns))
            tipe = gt(objek.find('.//dcc:model', ns))
            seri_issuer = gt(objek.find('.//dcc:identifications/dcc:identification/dcc:issuer', ns))
            seri_objek = gt(objek.find('.//dcc:identifications/dcc:identification/dcc:value', ns))

            id_lain_elements = objek.findall('.//dcc:identifications/dcc:identification/dcc:name/dcc:content', ns)
            id_lain = ', '.join([gt(id_lain_element) for id_lain_element in id_lain_elements if id_lain_element.text is not None])

            ws.cell(row=row, column=1, value=f"Objek {i}").font = Font(bold=True, size=12)
            row += 2
            ws.cell(row=row, column=1, value="Jenis alat atau objek")
            ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=5, value=jenis)
            row += 2
            ws.cell(row=row, column=1, value="Merek/pembuat")
            ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=5, value=merek)
            row += 2
            ws.cell(row=row, column=1, value="Tipe")
            ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=5, value=tipe)
            row += 2
            ws.cell(row=row, column=1, value="Identifikasi alat").font = Font(bold=True)
            row += 2
            ws.cell(row=row, column=1, value="Penerbit nomor seri")
            ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=5, value=seri_issuer)
            row += 2
            ws.cell(row=row, column=1, value="Nomor seri")
            ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=5, value=seri_objek)
            row += 2
            ws.cell(row=row, column=1, value="Identifikasi lain")
            ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=5, value=id_lain)
            row += 2

        #INFO LAB
        lab = root.find('.//dcc:calibrationLaboratory', ns)
        
        kode = gt(lab.find('dcc:calibrationLaboratoryCode', ns))
        nama_lab = gt(lab.find('dcc:contact/dcc:name/dcc:content', ns))
        email = gt(lab.find('dcc:contact/dcc:eMail', ns))
        telp = gt(lab.find('dcc:contact/dcc:phone', ns))
        link = gt(lab.find('dcc:contact/dcc:link', ns))
        jalan_lab = gt(lab.find('dcc:contact/dcc:location/dcc:street', ns))
        no_jalan_lab = gt(lab.find('dcc:contact/dcc:location/dcc:streetNo', ns))
        city_lab = gt(lab.find('dcc:contact/dcc:location/dcc:city', ns))
        state_lab = gt(lab.find('dcc:contact/dcc:location/dcc:state', ns))
        pos_lab = gt(lab.find('dcc:contact/dcc:location/dcc:postCode', ns))
        negara_lab = gt(lab.find('dcc:contact/dcc:location/dcc:countryCode', ns))

        ws.cell(row=row, column=1, value="Informasi Laboratorium Kalibrasi").font = Font(bold=True, size=14)
        row += 2
        ws.cell(row=row, column=1, value="Negara tempat kalibrasi")
        ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=5, value=negara_lab)
        row += 2
        ws.cell(row=row, column=1, value="Kode")
        ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=5, value=kode)
        row += 2
        ws.cell(row=row, column=1, value="Nama lab")
        ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=5, value=nama_lab)
        row += 2
        ws.cell(row=row, column=1, value="E-mail")
        ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=5, value=email)
        row += 2
        ws.cell(row=row, column=1, value="Nomor telepon")
        ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=5, value=telp)
        row += 2
        ws.cell(row=row, column=1, value="Link website")
        ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=5, value=link)
        row += 2
        ws.cell(row=row, column=1, value="Alamat").font = Font(bold=True)
        row += 2
        ws.cell(row=row, column=1, value="Jalan")
        ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=5, value=jalan_lab)
        row += 2
        ws.cell(row=row, column=1, value="Nomor jalan")
        ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=5, value=no_jalan_lab)
        row += 2
        ws.cell(row=row, column=1, value="Kecamatan/kabupaten/kota")
        ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=5, value=city_lab)
        row += 2
        ws.cell(row=row, column=1, value="Provinsi")
        ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=5, value=state_lab)
        row += 2
        ws.cell(row=row, column=1, value="Kode pos")
        ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=5, value=pos_lab)
        row += 2
        ws.cell(row=row, column=1, value="Negara")
        ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=5, value=negara_lab)
        row += 2

        #PENANGGUNG JAWAB
        resp = root.find('.//dcc:respPersons', ns)

        # Mengambil data penanggung jawab
        for person in resp.findall('dcc:respPerson', ns):
            # Ambil nama penanggung jawab
            nama_resp = person.find('dcc:person/dcc:name/dcc:content', ns).text if person.find('dcc:person/dcc:name/dcc:content', ns) is not None else 'N/A'
            nip_resp = person.find('dcc:description/dcc:content', ns).text if person.find('dcc:description/dcc:content', ns) is not None else 'N/A'
            
            # Ambil role jika ada, jika tidak, gunakan fallback
            role_element = person.find('dcc:role', ns)
            role = role_element.text if role_element is not None else 'N/A'
            
            # Cek apakah penanggung jawab utama atau bukan
            main_signer = person.find('dcc:mainSigner', ns)
            if main_signer is not None and main_signer.text == "true":
                role = "Direktur"
            
            # Menulis data ke Excel
            ws.cell(row=row, column=1, value=role).font = Font(bold=True)
            row += 2
            ws.cell(row=row, column=1, value="Nama")
            ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=5, value=nama_resp)
            row += 2
            ws.cell(row=row, column=1, value="NIP")
            ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=5, value=nip_resp)
            row += 2


        # OWNER INFO
        owner = root.find('.//dcc:customer', ns)

        nama_cust = gt(owner.find('dcc:name/dcc:content', ns))
        jalan_cust = gt(owner.find('dcc:location/dcc:street', ns))
        no_jalan_cust = gt(owner.find('dcc:location/dcc:streetNo', ns))
        city_cust = gt(owner.find('dcc:location/dcc:city', ns))
        state_cust = gt(owner.find('dcc:location/dcc:state', ns))
        pos_cust = gt(owner.find('dcc:location/dcc:postCode', ns))
        negara_cust = gt(owner.find('dcc:location/dcc:countryCode', ns))

        # Write owner (Pemilik Objek) data to Excel
        ws.cell(row=row, column=1, value="Identitas Pemilik Objek yang Dikalibrasi/Diukur").font = Font(bold=True, size=14)
        row += 2
        ws.cell(row=row, column=1, value="Nama")
        ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=5, value=nama_cust)
        row += 2
        ws.cell(row=row, column=1, value="Alamat").font = Font(bold=True)
        row += 2
        ws.cell(row=row, column=1, value="Jalan")
        ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=5, value=jalan_cust)
        row += 2
        ws.cell(row=row, column=1, value="Nomor jalan")
        ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=5, value=no_jalan_cust)
        row += 2
        ws.cell(row=row, column=1, value="Kecamatan/kabupaten/kota")
        ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=5, value=city_cust)
        row += 2
        ws.cell(row=row, column=1, value="Provinsi")
        ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=5, value=state_cust)
        row += 2
        ws.cell(row=row, column=1, value="Kode pos")
        ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=5, value=pos_cust)
        row += 2
        ws.cell(row=row, column=1, value="Negara")
        ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=5, value=negara_cust)
        row += 2
        
        # Metode Kalibrasi
        methods = root.findall('.//dcc:measurementResults/dcc:measurementResult/dcc:usedMethods/dcc:usedMethod', ns)
        ws.cell(row=row, column=1, value="Metode Kalibrasi").font = Font(bold=True, size=14)
        row += 2

        # Assuming methods are provided as a list, iterate over each method
        for i, method in enumerate(methods, start=1):
            
            method_name_elements = method.findall('dcc:name/dcc:content', ns)
            method_name = ', '.join([gt(method_name_element) for method_name_element in method_name_elements if method_name_element.text])

            method_desc_elements = method.findall('dcc:description/dcc:content', ns)
            method_desc = ', '.join([gt(method_desc_element) for method_desc_element in method_desc_elements if method_desc_element.text])

            norm = gt(method.find('dcc:norm', ns))

            ws.cell(row=row, column=1, value=f"Metode {i}").font = Font(bold=True, size=12)
            row += 2
            ws.cell(row=row, column=1, value="Nama")
            ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=5, value=method_name)
            row += 2
            ws.cell(row=row, column=1, value="Norma")
            ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=5, value=norm)
            row += 2
            ws.cell(row=row, column=1, value="Deskripsi")
            ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=5, value=method_desc)
            row += 2
            
            # RUMUS
            formula_elem = method.find('dcc:description/dcc:formula/dcc:latex', ns)
            latex_raw = gt(formula_elem) if formula_elem is not None else ""

            ws.cell(row=row, column=1, value="Rumus")
            ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')

            if latex_raw:
                try:
                    latex_buf = render_latex_to_excel_image(latex_raw)
                    latex_img = ExcelImage(latex_buf)
                    
                    # Atur ukuran maksimum di Excel
                    max_excel_width = 400  # pixel
                    if latex_img.width > max_excel_width:
                        ratio = max_excel_width / latex_img.width
                        latex_img.width = max_excel_width
                        latex_img.height = int(latex_img.height * ratio)
                    
                    # Hitung tinggi baris yang diperlukan (1.2 = faktor padding)
                    row_height = max(20, int(latex_img.height * 1.2))
                    ws.row_dimensions[row].height = row_height
                    
                    ws.add_image(latex_img, f"E{row}")
                    row += int(row_height / 25) + 1
                    
                except Exception as e:
                    logging.error(f"[ERROR] Gagal render rumus: {e}")
                    ws.cell(row=row, column=5, value="[Rumus gagal ditampilkan]")
                    row += 2
            else:
                ws.cell(row=row, column=5, value="-")
                row += 2
                
            
            # GAMBAR
            description_elem = method.find('dcc:description', ns)
            file_elem = description_elem.find('dcc:file', ns) if description_elem is not None else None

            if file_elem is not None:
                mime_type = gt(file_elem.find('dcc:mimeType', ns))
                caption = gt(file_elem.find('dcc:caption', ns))
                data_base64_elem = file_elem.find('dcc:dataBase64', ns)
                ws.cell(row=row, column=1, value="Gambar")
                ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
                row += 1

                if data_base64_elem is not None and data_base64_elem.text:
                    base64_str = data_base64_elem.text

                    try:
                        row_used = add_image_to_worksheet(ws, base64_str, mime_type, caption, row, 'E')
                        row += row_used + 5 
                    except Exception as e:
                        logging.error(f"[ERROR] Gagal menambahkan gambar di metode kalibrasi: {e}")
                        ws.cell(row=row, column=5, value="[Gambar gagal dimuat]")
                        row += 2
                else:
                    ws.cell(row=row, column=5, value="[Base64 kosong]")
                    row += 2
            else:
                ws.cell(row=row, column=5, value="[Tidak ada gambar]")
                row += 2


        # Standar atau Alat Pengukuran
        equipments = root.findall('.//dcc:measuringEquipments/dcc:measuringEquipment', ns)
        ws.cell(row=row, column=1, value="Standar atau Alat Pengukuran").font = Font(bold=True, size=14)
        row += 2

        # Iterate through each equipment (measuring equipment)
        for i, equip in enumerate(equipments, start=1):
            # Handle 'nama_alat' (equipment name) in multiple languages
            nama_alat_elements = equip.findall('dcc:name/dcc:content', ns)
            nama_alat = ', '.join([gt(nama_alat_element) for nama_alat_element in nama_alat_elements if nama_alat_element.text])

            # Handle 'model' (manufacturer model) in multiple languages
            manuf_model_elements = equip.findall('dcc:manufacturer/dcc:name/dcc:content', ns)
            model = ', '.join([gt(manuf_model_element) for manuf_model_element in manuf_model_elements if manuf_model_element.text])

            # Extract serial number (seri_measuring)
            seri_measuring = gt(equip.find('dcc:identifications/dcc:identification/dcc:value', ns))

            # Handle 'manuf_model' (manufacturer model in identification) in multiple languages
            manuf_model_id_elements = equip.findall('dcc:identifications/dcc:identification/dcc:name/dcc:content', ns)
            manuf_model = ', '.join([gt(manuf_model_id_element) for manuf_model_id_element in manuf_model_id_elements if manuf_model_id_element.text])

           
            ws.cell(row=row, column=1, value=f"Alat {i}").font = Font(bold=True, size=12)
            row += 2
            ws.cell(row=row, column=1, value="Nama")
            ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=5, value=nama_alat)
            row += 2
            ws.cell(row=row, column=1, value="Merek dan model")
            ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=5, value=model)
            row += 2
            ws.cell(row=row, column=1, value="Nomor seri")
            ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=5, value=seri_measuring)
            row += 2
            ws.cell(row=row, column=1, value="Merek dan model dalam identifikasi")
            ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=5, value=manuf_model)
            row += 2

        # Kondisi Lingkungan
        conditions = root.findall('.//dcc:influenceConditions/dcc:influenceCondition', ns)
        ws.cell(row=row, column=1, value="Kondisi Lingkungan").font = Font(bold=True, size=14)
        row += 2

        # Iterate over each condition (e.g., temperature, humidity, or both)
        for i, kondisi in enumerate(conditions, start=1):
            # Extract the condition name (e.g., temperature or humidity) in multiple languages
            parameter_elements = kondisi.findall('dcc:name/dcc:content', ns)
            parameter = ', '.join([gt(parameter_element) for parameter_element in parameter_elements if parameter_element.text])

            # Extract the condition description in multiple languages
            kondisi_desc_elements = kondisi.findall('dcc:description/dcc:content', ns)
            kondisi_desc = ', '.join([gt(kondisi_desc_element) for kondisi_desc_element in kondisi_desc_elements if kondisi_desc_element.text])

            
            ws.cell(row=row, column=1, value=f"Parameter {i}").font = Font(bold=True, size=12)
            row += 2
            ws.cell(row=row, column=1, value="Parameter lingkungan")
            ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=5, value=parameter)
            row += 2
            ws.cell(row=row, column=1, value="Deskripsi")
            ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=5, value=kondisi_desc)
            row += 2

            # Extracting variables (values) for this condition
            variables = kondisi.findall('dcc:data/dcc:quantity', ns)
            for var in variables:
                
                variabel = gt(var.find('dcc:name/dcc:content', ns))
               
                kondisi_value = gt(var.find('si:real/si:value', ns))
         
                kondisi_unit = gt(var.find('si:real/si:unit', ns))

                # Convert unit to a more readable form
                kondisi_unit_converted = convert_unit(kondisi_unit)

                
                ws.cell(row=row, column=1, value=variabel)
                ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
                ws.cell(row=row, column=5, value=kondisi_value)
                ws.cell(row=row, column=6, value=kondisi_unit_converted)
                row += 2
 

        # Hasil Kalibrasi
        ws.cell(row=row, column=1, value="Hasil Kalibrasi").font = Font(bold=True, size=14)
        row += 2

        results = root.findall('.//dcc:results/dcc:result', ns)

        for hasil in results:
            quantities = hasil.findall('dcc:data/dcc:list/dcc:quantity', ns)
            quantity_names = []
            for q in quantities:
                content = q.find('dcc:name/dcc:content[@lang="en"]', ns)
                if content is None:
                    content = q.find('dcc:name/dcc:content', ns)
                quantity_names.append(gt(content) if content is not None else "No Name")

            all_quantity_data = []
            max_rows = 0

            for q in quantities:
                hybrid = q.find('si:hybrid', ns)
                quantity_data = []

                if hybrid is not None:
                    real_lists = hybrid.findall('si:realListXMLList', ns)
                else:
                    real_lists = q.findall('si:realListXMLList', ns)

                for rl in real_lists:
                    value_elem = rl.find('si:valueXMLList', ns)
                    unit_elem = rl.find('si:unitXMLList', ns)

                    values = value_elem.text.strip().split() if value_elem is not None and value_elem.text else []
                    units = unit_elem.text.strip().split() if unit_elem is not None and unit_elem.text else []

                    if len(units) == 1:
                        unit = units[0]
                        quantity_data.append([(v, convert_unit(unit)) for v in values])
                    else:
                        quantity_data.append([(v, convert_unit(unit)) for v, unit in zip(values, units)])

                all_quantity_data.append(quantity_data)
                if quantity_data:
                    max_rows = max(max_rows, len(quantity_data[0]))

            # Ambil uncertainty dari salah satu quantity
            uncertainty_values = []
            for q in quantities:
                mu_elem = q.find('.//si:valueExpandedMUXMLList', ns)
                if mu_elem is not None and mu_elem.text:
                    uncertainty_values = mu_elem.text.strip().split()
                    break  # pakai hanya satu uncertainty list

            # Nama tabel
            title_elem = hasil.find('dcc:name/dcc:content[@lang="en"]', ns)
            if title_elem is None:
                title_elem = hasil.find('dcc:name/dcc:content', ns)
            title = gt(title_elem) if title_elem is not None else "No Title"

            # Hitung total kolom termasuk uncertainty (2 kolom extra)
            total_columns = sum(len(q_data) * 2 for q_data in all_quantity_data) + (2 if uncertainty_values else 0)

            # Judul tabel
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_columns)
            table_name = ws.cell(row=row, column=1, value=title)
            table_name.font = Font(bold=True)
            apply_borders(ws, row, row, 1, total_columns)
            row += 1

            col = 1
            header_range = []

            # Header untuk kolom quantity
            for q_name, q_data in zip(quantity_names, all_quantity_data):
                span = len(q_data)
                header_cell = ws.cell(row=row, column=col, value=q_name)
                header_cell.font = Font(bold=True)
                header_cell.alignment = Alignment(horizontal="center")
                header_range.append((col, col + span * 2 - 1))
                col += span * 2

            # Header untuk uncertainty
            if uncertainty_values:
                ws.cell(row=row, column=col, value="Ketidakpastian / Uncertainty").font = Font(bold=True)
                ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
                header_range.append((col, col + 1))
                col += 2

            for start_col, end_col in header_range:
                if start_col > end_col:
                    start_col, end_col = end_col, start_col
                ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)

            apply_borders(ws, row, row, 1, total_columns)
            row += 1

            # Data baris per baris
            for i in range(max_rows):
                col = 1
                last_unit = ""  # untuk digunakan di kolom uncertainty
                for q_data in all_quantity_data:
                    for pair_set in q_data:
                        if i < len(pair_set):
                            val, unit = pair_set[i]
                            last_unit = unit
                            ws.cell(row=row, column=col, value=round(float(val), 5))
                            ws.cell(row=row, column=col + 1, value=unit)
                        col += 2

                # Isi kolom uncertainty
                if i < len(uncertainty_values):
                    ws.cell(row=row, column=col, value=round(float(uncertainty_values[i]), 5))
                    ws.cell(row=row, column=col + 1, value=last_unit)
                row += 1

            apply_borders(ws, row - max_rows, row - 1, 1, total_columns)
            row += 1

            #Informasi Ketidakpastian
            coverage_factor = coverage_prob = distribution = None
            for q in quantities:
                coverage_factor_elem = q.find('.//si:coverageFactorXMLList', ns)
                coverage_prob_elem = q.find('.//si:coverageProbabilityXMLList', ns)
                distribution_elem = q.find('.//si:distributionXMLList', ns)

                if coverage_factor_elem is not None and coverage_factor_elem.text:
                    coverage_factor = coverage_factor_elem.text.strip()
                if coverage_prob_elem is not None and coverage_prob_elem.text:
                    coverage_prob = coverage_prob_elem.text.strip()
                if distribution_elem is not None and distribution_elem.text:
                    distribution = distribution_elem.text.strip()

                # Gunakan yang pertama ditemukan
                if coverage_factor or coverage_prob or distribution:
                    break

            # Tulis ke worksheet jika ada
            if coverage_factor:
                ws.cell(row=row, column=1, value="Faktor Cakupan")
                ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
                ws.cell(row=row, column=5, value=coverage_factor)
                row += 1
            if coverage_prob:
                ws.cell(row=row, column=1, value="Probabilitas Cakupan")
                ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
                ws.cell(row=row, column=5, value=coverage_prob)
                row += 1
            if distribution:
                ws.cell(row=row, column=1, value="Distribusi")
                ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
                ws.cell(row=row, column=5, value=distribution)
                row += 1
            row += 3

        #STATEMENT/PERNYATAAN
        statements = root.findall('.//dcc:statements/dcc:statement', ns)
        ws.cell(row=row, column=1, value="Pernyataan").font = Font(bold=True, size=14)
        row += 2

        for i, pernyataan in enumerate(statements, start=1):
            declaration = pernyataan.find('dcc:declaration', ns)
            
            statement_elements = pernyataan.findall('dcc:declaration/dcc:content', ns)
            statement_text = ', '.join([gt(statement_element) for statement_element in statement_elements if statement_element.text])

            ws.cell(row=row, column=1, value=f"Pernyataan {i}").font = Font(bold=True, size=12)
            row += 1
            ws.cell(row=row, column=1, value=statement_text)
            row += 2
            
            # RUMUS
            formula_elem = declaration.find('dcc:formula/dcc:latex', ns) if declaration is not None else None
            latex_raw = gt(formula_elem) if formula_elem is not None else ""

            ws.cell(row=row, column=1, value="Rumus")
            ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')

            if latex_raw:
                try:
                    latex_buf = render_latex_to_excel_image(latex_raw)
                    latex_img = ExcelImage(latex_buf)

                    max_excel_width = 120
                    if latex_img.width > max_excel_width:
                        ratio = max_excel_width / latex_img.width
                        latex_img.width = max_excel_width
                        latex_img.height = int(latex_img.height * ratio)

                    ws.row_dimensions[row].height = latex_img.height * 0.75

                    img_anchor = f"E{row}"
                    ws.add_image(latex_img, img_anchor)

                    row += 1
                    row += 5

                except Exception as e:
                    logging.error(f"[ERROR] Gagal render rumus di pernyataan: {e}")
                    ws.cell(row=row, column=5, value="[Rumus gagal ditampilkan]")
                    row += 2
            else:
                ws.cell(row=row, column=5, value="-")
                row += 2
            
            
            # GAMBAR
            file_elem = declaration.find('dcc:file', ns) if declaration is not None else None

            if file_elem is not None:
                mime_type = gt(file_elem.find('dcc:mimeType', ns))
                caption = gt(file_elem.find('dcc:caption', ns))
                data_base64_elem = file_elem.find('dcc:dataBase64', ns)

                ws.cell(row=row, column=1, value="Gambar")
                ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal='center')
                row += 1

                if data_base64_elem is not None and data_base64_elem.text:
                    base64_str = data_base64_elem.text
                    try:
                        row_used = add_image_to_worksheet(ws, base64_str, mime_type, caption, row, 'E')
                        row += row_used + 5
                    except Exception as e:
                        logging.error(f"[ERROR] Gagal menambahkan gambar di pernyataan: {e}")
                        ws.cell(row=row, column=5, value="[Gambar gagal dimuat]")
                        row += 2
                else:
                    ws.cell(row=row, column=5, value="[Base64 kosong]")
                    row += 2
            else:
                ws.cell(row=row, column=5, value="[Tidak ada gambar]")
                row += 2
        row += 3
                
                
        #COMMENT
        comment_elem = root.find('.//dcc:comment', ns)
        if comment_elem is not None:
            ws.cell(row=row, column=1, value="Komentar").font = Font(bold=True, size=14)
            row += 2

            # Nama komentar
            name_elem = comment_elem.find('dcc:name/dcc:content', ns)
            comment_name = gt(name_elem) if name_elem is not None else "-"

            ws.cell(row=row, column=1, value="Nama")
            ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=5, value=comment_name)
            row += 2

            # Deskripsi komentar (multi bahasa)
            desc_elems = comment_elem.findall('dcc:description/dcc:content', ns)
            desc_texts = [gt(de) for de in desc_elems if de is not None and de.text]

            ws.cell(row=row, column=1, value="Deskripsi")
            ws.cell(row=row, column=4, value=":").alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=5, value=', '.join(desc_texts) if desc_texts else "-")
            row += 3  # Jarak ke bagian setelahnya

                

        # Save the Excel file
        excel_filename = os.path.basename(xml_file_path) + ".xlsx"
        excel_path = os.path.join(os.path.dirname(xml_file_path), excel_filename)
        wb.save(excel_path)

        return excel_path

    except Exception as e:
        raise Exception(f"Error converting XML to Excel: {str(e)}")
